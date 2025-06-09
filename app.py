from flask import Flask, render_template, request, redirect, url_for, flash, session, send_from_directory
from werkzeug.security import generate_password_hash, check_password_hash
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
import openpyxl
import os
from datetime import datetime, date 
import re
import glob
import math
import json 
from functools import wraps # Pastikan impor ini ada

app = Flask(__name__)
app.secret_key = 'kunci_rahasia_super_aman_milik_abu_zayd_123!@#' 

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login_route' 
login_manager.login_message = "Anda harus login untuk mengakses halaman ini."
login_manager.login_message_category = "info"

USERS_FILE = "users.json"
DEFAULT_EXCEL_FILE = "data_manifest_default.xlsx"
EXCEL_STORAGE_DIR = "excel_files" 
if not os.path.exists(EXCEL_STORAGE_DIR):
    os.makedirs(EXCEL_STORAGE_DIR)

headers_excel = [
    "FULL NAME", "SEX", "PLACE OF BIRTH", "DATE OF BIRTH", "AGE",
    "PASPORT NO", "DATE OF ISSUE", "DATE OF EXPIRY", "ISSUING OFFICE",
    "MENINGITIS EXPIRY", "NAMA SESUAI VAKSIN", "NOMOR BPJS", "MAHRAM",
    "NAMA AYAH", "NIK", "ALAMAT", "PROVINSI", "KABUPATEN", "KECAMATAN",
    "KELURAHAN", "NO. HP", "STATUS PERNIKAHAN", "PENDIDIKAN", "PEKERJAAN",
    "TYPE OF ROOM", "BRANCH"
]

# --- Fungsi Helper Pengguna ---
def load_users():
    if not os.path.exists(USERS_FILE):
        save_users({}) 
        return {}
    try:
        with open(USERS_FILE, 'r') as f:
            users_data = json.load(f)
            return {str(k): v for k, v in users_data.items()}
    except json.JSONDecodeError:
        print(f"Peringatan: File {USERS_FILE} korup atau kosong. Membuat file baru.")
        save_users({}) 
        return {} 
    except Exception as e:
        print(f"Error saat memuat pengguna: {e}")
        return {}

def save_users(users_data):
    try:
        with open(USERS_FILE, 'w') as f:
            json.dump(users_data, f, indent=4)
    except Exception as e:
        print(f"Error saat menyimpan pengguna: {e}")

class User(UserMixin):
    def __init__(self, id, username, password_hash, role="user", status="pending_approval"):
        self.id = str(id) 
        self.username = username
        self.password = password_hash 
        self.role = role 
        self.status = status 

    @staticmethod
    def get(user_id):
        users = load_users()
        user_data = users.get(str(user_id)) 
        if user_data and 'username' in user_data and 'password' in user_data:
            return User(id=str(user_id), username=user_data['username'], 
                        password_hash=user_data['password'], 
                        role=user_data.get('role', 'user'),
                        status=user_data.get('status', 'pending_approval'))
        return None

    @staticmethod
    def find_by_username(username):
        users = load_users()
        for user_id, user_data in users.items(): 
            if user_data.get('username') == username and 'password' in user_data: 
                return User(id=str(user_id), username=user_data['username'], 
                            password_hash=user_data['password'], 
                            role=user_data.get('role', 'user'),
                            status=user_data.get('status', 'pending_approval'))
        return None

@login_manager.user_loader
def load_user(user_id):
    if user_id is None: return None
    user = User.get(str(user_id))
    if user and (user.status == 'active' or user.status == 'superuser'):
        return user
    return None 

# === PASTIKAN DEFINISI DECORATOR INI ADA SEBELUM DIGUNAKAN ===
def superuser_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or getattr(current_user, 'role', 'user') != 'superuser':
            flash('Akses ditolak: Anda harus menjadi superuser untuk mengakses fitur ini.', 'danger')
            return redirect(url_for('main_page_route')) 
        return f(*args, **kwargs)
    return decorated_function
# ============================================================

# --- Fungsi Helper Aplikasi Manifest ---
def get_file_path(filename): return os.path.join(EXCEL_STORAGE_DIR, filename)
def is_valid_date_format(d, f="%Y-%m-%d"): return True if not d else Falsify(d,f)
def Falsify(d,f):
    try: datetime.strptime(d,f); return True
    except ValueError: return False
def add_months_to_date(sourcedate, months):
    month = sourcedate.month - 1 + months; year = sourcedate.year + month // 12
    month = month % 12 + 1
    day = min(sourcedate.day, [31,29 if year%4==0 and (year%100!=0 or year%400==0) else 28,31,30,31,30,31,31,30,31,30,31][month-1])
    return date(year, month, day)
def calculate_age_from_dob_string(dob_string):
    if not dob_string: return "0"
    formats_to_try = ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%b-%Y", "%d-%b-%y"]
    parsed_date = None
    for fmt in formats_to_try:
        try:
            if fmt.endswith("%y") and not fmt.endswith("-%b-%y") and fmt != "%Y-%m-%d":
                temp_date = datetime.strptime(dob_string, fmt); year_short = temp_date.year % 100
                current_year_short = datetime.today().year % 100
                parsed_date = datetime(year_short + (1900 if year_short > (current_year_short + 15) and year_short <= 99 else 2000), temp_date.month, temp_date.day).date()
            else: parsed_date = datetime.strptime(dob_string, fmt).date()
            break
        except ValueError: pass
    if parsed_date:
        today = date.today(); age = today.year - parsed_date.year - ((today.month, today.day) < (parsed_date.month, parsed_date.day))
        return str(age)
    return "N/A"
def initialize_excel_file(filename_to_initialize):
    filepath = get_file_path(filename_to_initialize)
    if not os.path.exists(filepath):
        workbook = openpyxl.Workbook(); sheet = workbook.active; sheet.title = "Manifest Data"
        sheet.append(headers_excel); workbook.save(filepath)
        print(f"File {filename_to_initialize} dibuat di {EXCEL_STORAGE_DIR}.")
    else:
        try:
            workbook = openpyxl.load_workbook(filepath)
            if "Manifest Data" not in workbook.sheetnames:
                sheet = workbook.create_sheet("Manifest Data"); sheet.append(headers_excel); workbook.save(filepath)
            elif workbook["Manifest Data"].max_row == 0 or (workbook["Manifest Data"].max_row == 1 and all(c.value is None for c in workbook["Manifest Data"][1])):
                sheet = workbook["Manifest Data"]
                if sheet.max_row > 0: sheet.delete_rows(1, sheet.max_row)
                sheet.append(headers_excel); workbook.save(filepath)
        except Exception as e: print(f"Error inisialisasi sheet {filename_to_initialize}: {e}")
def list_excel_files(): return sorted([os.path.basename(f) for f in glob.glob(os.path.join(EXCEL_STORAGE_DIR, "*.xlsx"))])
def get_all_manifest_data(target_filename, page=1, per_page=10):
    all_items_with_ids = []
    filepath = get_file_path(target_filename)
    empty_pagination = {'items': [], 'page': page, 'per_page': per_page, 'total_items': 0, 'total_pages': 0, 'has_prev': False, 'prev_num': None, 'has_next': False, 'next_num': None}
    if not os.path.exists(filepath): return empty_pagination
    try:
        workbook = openpyxl.load_workbook(filepath)
        if "Manifest Data" in workbook.sheetnames:
            sheet = workbook["Manifest Data"]
            if not isinstance(sheet.max_row, int): return empty_pagination
            for row_index in range(2, sheet.max_row + 1):
                current_row_values = [str(sheet.cell(row=row_index, column=col_index).value or "") for col_index in range(1, len(headers_excel) + 1)]
                if any(val != "" for val in current_row_values):
                    all_items_with_ids.append({"excel_row_num": row_index, "values": current_row_values})
    except Exception as e:
        print(f"Error baca Excel {target_filename}: {e}"); flash(f'Gagal muat data dari {target_filename}: {e}', 'danger')
        return empty_pagination
    total_items = len(all_items_with_ids); total_pages = math.ceil(total_items / per_page) if per_page > 0 else 0
    page = max(1, min(page, total_pages if total_pages > 0 else 1))
    start_index = (page - 1) * per_page; end_index = start_index + per_page
    paginated_items = all_items_with_ids[start_index:end_index]
    return {'items': paginated_items, 'page': page, 'per_page': per_page, 'total_items': total_items, 'total_pages': total_pages, 'has_prev': page > 1, 'prev_num': page - 1 if page > 1 else None, 'has_next': page < total_pages, 'next_num': page + 1 if page < total_pages else None}
def process_and_validate_form(form_request):
    form_data = {key: form_request.get(key, '').strip() if isinstance(form_request.get(key), str) else form_request.get(key, '') for key in form_request.keys()}
    errors = [];
    if not form_data.get("form_full_name"): errors.append("Nama Lengkap wajib diisi.")
    sex_value = form_data.get("form_sex")
    if sex_value and sex_value not in ['M', 'F']: errors.append("Jenis Kelamin tidak valid.")
    dob_str = form_data.get("form_date_of_birth")
    age_calculation_result = calculate_age_from_dob_string(dob_str) 
    if dob_str and age_calculation_result == "N/A": errors.append("Format Tanggal Lahir tidak valid.")
    nik_value = form_data.get("form_nik")
    if nik_value and not (nik_value.isdigit() and len(nik_value) == 16): errors.append("NIK harus 16 digit angka.")
    no_hp_value = form_data.get("form_no_hp")
    if no_hp_value and not re.match(r"^\+?\d{10,15}$", no_hp_value): errors.append("Format Nomor HP tidak valid.")
    no_bpjs_value = form_data.get("form_nomor_bpjs")
    if no_bpjs_value and not (no_bpjs_value.isdigit() and len(no_bpjs_value) == 13): errors.append("Nomor BPJS harus 13 digit angka.")
    date_fields_to_check = {"form_date_of_issue": "Tg Terbit Paspor", "form_date_of_expiry": "Tg Kedaluwarsa Paspor", "form_meningitis_expiry": "Tg Kedaluwarsa Vaksin Meningitis"}
    parsed_dates = {} 
    for field_name, label in date_fields_to_check.items():
        date_str_value = form_data.get(field_name)
        if date_str_value:
            if not is_valid_date_format(date_str_value): errors.append(f"Format {label} tidak valid (YYYY-MM-DD).")
            else:
                try: parsed_dates[field_name] = datetime.strptime(date_str_value, "%Y-%m-%d").date()
                except ValueError: errors.append(f"Error parsing internal untuk {label}.")
    tgl_terbit = parsed_dates.get("form_date_of_issue"); tgl_kedaluwarsa_paspor = parsed_dates.get("form_date_of_expiry")
    if tgl_kedaluwarsa_paspor: 
        today = date.today()
        if tgl_kedaluwarsa_paspor < today: errors.append(f"Paspor sudah kedaluwarsa pada {tgl_kedaluwarsa_paspor.strftime('%d-%b-%Y')}.")
        else: 
            min_acceptable_expiry = add_months_to_date(today, 8)
            if tgl_kedaluwarsa_paspor < min_acceptable_expiry:
                form_data['passport_expiry_warning'] = f"Peringatan: Masa berlaku paspor ({tgl_kedaluwarsa_paspor.strftime('%d-%b-%Y')}) kurang dari 8 bulan dari sekarang."
        if tgl_terbit: 
            if tgl_terbit >= tgl_kedaluwarsa_paspor:
                if not any("Paspor sudah kedaluwarsa" in e for e in errors): errors.append("Tg kedaluwarsa paspor harus setelah tg terbit.")
        elif not tgl_terbit: errors.append("Jika tanggal kedaluwarsa paspor diisi, tanggal penerbitan juga harus diisi.")
    return form_data, errors, age_calculation_result

@app.route('/')
@login_required 
def main_page_route():
    active_file = session.get('active_excel_file')
    page = request.args.get('page', 1, type=int) 
    per_page = 10 
    pagination_data = {'items': [], 'page': page, 'per_page': per_page, 'total_items': 0, 'total_pages': 0, 'has_prev': False, 'prev_num': None, 'has_next': False, 'next_num': None}
    form_data_to_render = session.pop('form_data_on_error', None) 
    initial_tab = session.pop('initial_tab_on_redirect', 'beranda') 
    all_users_list = []
    pending_approval_users = []
    if current_user.is_authenticated and current_user.role == 'superuser':
        users_dict = load_users()
        for user_id, user_data in users_dict.items():
            user_info = {'id': user_id, 'username': user_data.get('username'), 'role': user_data.get('role'), 'status': user_data.get('status')}
            if user_data.get('status') == 'pending_approval':
                pending_approval_users.append(user_info)
            elif user_data.get('status') in ['active', 'superuser']: 
                all_users_list.append(user_info)
    if active_file:
        initialize_excel_file(active_file) 
        pagination_data = get_all_manifest_data(active_file, page=page, per_page=per_page)
    else: 
        initialize_excel_file(DEFAULT_EXCEL_FILE)
        pagination_data = get_all_manifest_data(DEFAULT_EXCEL_FILE, page=page, per_page=per_page)
        if not active_file and pagination_data.get('total_items', 0) > 0:
            flash(f"Menampilkan data dari file default: {DEFAULT_EXCEL_FILE}. Pilih atau buat file di Beranda untuk mulai bekerja.", "info")
    edit_mode_data = session.get('edit_mode_data')
    if edit_mode_data:
        form_data_to_render = edit_mode_data.get('form_values')
        if form_data_to_render is None: form_data_to_render = {}
        form_data_to_render['editing_excel_row_num'] = edit_mode_data.get('excel_row_num')
        form_data_to_render['editing_filename'] = edit_mode_data.get('filename')
        initial_tab = 'input' 
    available_files = list_excel_files()
    if not available_files and not os.path.exists(get_file_path(DEFAULT_EXCEL_FILE)):
        initialize_excel_file(DEFAULT_EXCEL_FILE)
        available_files = list_excel_files()
    return render_template('index.html', 
                           user=current_user, 
                           headers=headers_excel, 
                           pagination=pagination_data, 
                           form_data=form_data_to_render,
                           available_files=available_files,
                           default_excel_file=DEFAULT_EXCEL_FILE,
                           active_excel_file=active_file,
                           initial_tab=initial_tab,
                           all_users=all_users_list, 
                           pending_users=pending_approval_users)

@app.route('/set_active_file', methods=['POST'])
@login_required
def set_active_file_route():
    file_choice = request.form.get('file_choice')
    new_excel_filename_input = request.form.get('new_excel_filename', '').strip()
    existing_excel_filename_input = request.form.get('existing_excel_filename', '')
    active_file = None; error_file_selection = None
    if file_choice == 'new_file':
        if not new_excel_filename_input: error_file_selection = "Nama file baru tidak boleh kosong."
        else:
            if not new_excel_filename_input.lower().endswith('.xlsx'): new_excel_filename_input += '.xlsx'
            if any(c in new_excel_filename_input for c in ['/', '\\', '..']): error_file_selection = "Nama file baru mengandung karakter yang tidak valid."
            else: active_file = new_excel_filename_input; initialize_excel_file(active_file) 
    elif file_choice == 'existing_file':
        if not existing_excel_filename_input: error_file_selection = "Pilih file Excel yang sudah ada."
        elif not os.path.exists(get_file_path(existing_excel_filename_input)): error_file_selection = f"File '{existing_excel_filename_input}' tidak ditemukan."
        else: active_file = existing_excel_filename_input
    else: error_file_selection = "Pilihan file tidak valid."
    if error_file_selection: flash(error_file_selection, 'danger'); return redirect(url_for('main_page_route'))
    if active_file:
        session['active_excel_file'] = active_file; session.pop('edit_mode_data', None) 
        flash(f"File aktif diatur ke: {active_file}", "info"); session['initial_tab_on_redirect'] = 'input' 
    else: flash("Gagal mengatur file aktif.", "danger")
    return redirect(url_for('main_page_route', page=1))

@app.route('/submit_entry', methods=['POST'])
@login_required
def submit_entry_route():
    active_file = session.get('active_excel_file')
    if not active_file: flash("Sesi file aktif tidak ditemukan.", "danger"); return redirect(url_for('main_page_route'))
    form_data_input = request.form.to_dict()
    processed_form_data, errors, age_calculation_result = process_and_validate_form(request.form)
    passport_warning = processed_form_data.pop('passport_expiry_warning', None)
    if errors:
        for error_msg in errors: flash(error_msg, 'danger')
        if passport_warning: flash(passport_warning, 'warning') 
        session['form_data_on_error'] = form_data_input 
        session['initial_tab_on_redirect'] = 'input'
        return redirect(url_for('main_page_route'))
    if passport_warning: flash(passport_warning, 'warning')
    age_to_save = age_calculation_result
    if age_calculation_result == "N/A" and processed_form_data.get("form_date_of_birth"): age_to_save = "N/A" 
    elif not processed_form_data.get("form_date_of_birth"): age_to_save = "0"
    data_to_map = { "FULL NAME": processed_form_data.get("form_full_name"), "SEX": processed_form_data.get("form_sex"), "PLACE OF BIRTH": processed_form_data.get("form_place_of_birth"), "DATE OF BIRTH": processed_form_data.get("form_date_of_birth"), "AGE": age_to_save, "PASPORT NO": processed_form_data.get("form_pasport_no"), "DATE OF ISSUE": processed_form_data.get("form_date_of_issue"), "DATE OF EXPIRY": processed_form_data.get("form_date_of_expiry"), "ISSUING OFFICE": processed_form_data.get("form_issuing_office"), "MENINGITIS EXPIRY": processed_form_data.get("form_meningitis_expiry"), "NAMA SESUAI VAKSIN": processed_form_data.get("form_nama_sesuai_vaksin"), "NOMOR BPJS": processed_form_data.get("form_nomor_bpjs"), "MAHRAM": processed_form_data.get("form_mahram"), "NAMA AYAH": processed_form_data.get("form_nama_ayah"), "NIK": processed_form_data.get("form_nik"), "ALAMAT": processed_form_data.get("form_alamat"), "PROVINSI": processed_form_data.get("form_provinsi"), "KABUPATEN": processed_form_data.get("form_kabupaten"), "KECAMATAN": processed_form_data.get("form_kecamatan"), "KELURAHAN": processed_form_data.get("form_kelurahan"), "NO. HP": processed_form_data.get("form_no_hp"), "STATUS PERNIKAHAN": processed_form_data.get("form_status_pernikahan"), "PENDIDIKAN": processed_form_data.get("form_pendidikan"), "PEKERJAAN": processed_form_data.get("form_pekerjaan"), "TYPE OF ROOM": processed_form_data.get("form_type_of_room"), "BRANCH": processed_form_data.get("form_branch")}
    row_data_to_save = [data_to_map.get(header, "") for header in headers_excel]
    try:
        filepath = get_file_path(active_file)
        workbook = openpyxl.load_workbook(filepath) 
        sheet = workbook["Manifest Data"] if "Manifest Data" in workbook.sheetnames else workbook.create_sheet("Manifest Data", 0)
        if sheet.max_row == 0 or (sheet.max_row == 1 and all(c.value is None for c in sheet[1])): 
            if sheet.max_row > 0 : sheet.delete_rows(1, sheet.max_row)
            sheet.append(headers_excel)
        sheet.append(row_data_to_save)
        workbook.save(filepath) 
        flash(f'Data manifest berhasil disimpan ke file {active_file}!', 'success')
        session.pop('form_data_on_error', None)
    except Exception as e:
        flash(f'Terjadi error saat menyimpan data ke {active_file}: {e}', 'danger')
        print(f"Error menyimpan ke Excel {active_file}: {e}")
    session['initial_tab_on_redirect'] = 'input'
    try:
        workbook = openpyxl.load_workbook(get_file_path(active_file))
        sheet = workbook["Manifest Data"]
        total_data_rows = sheet.max_row -1 
        last_page = math.ceil(total_data_rows / 10) if total_data_rows > 0 else 1
    except: last_page = 1 
    return redirect(url_for('main_page_route', page=last_page, tab='preview'))

@app.route('/edit_entry/<string:filename>/<int:excel_row_num>', methods=['GET'])
@login_required
def edit_entry_route(filename, excel_row_num):
    if '/' in filename or '\\' in filename or '..' in filename or not os.path.exists(get_file_path(filename)):
        flash("Nama file tidak valid atau tidak ditemukan.", "danger"); return redirect(url_for('main_page_route'))
    session['active_excel_file'] = filename 
    filepath = get_file_path(filename); form_values_for_edit = {}
    try:
        workbook = openpyxl.load_workbook(filepath); sheet = workbook["Manifest Data"]
        if 2 <= excel_row_num <= sheet.max_row:
            row_data = [str(sheet.cell(row=excel_row_num, column=col_idx).value or "") for col_idx in range(1, len(headers_excel) + 1)]
            form_field_mapping_edit = { "form_full_name": "FULL NAME", "form_sex": "SEX", "form_place_of_birth": "PLACE OF BIRTH", "form_date_of_birth": "DATE OF BIRTH", "form_pasport_no": "PASPORT NO", "form_date_of_issue": "DATE OF ISSUE", "form_date_of_expiry": "DATE OF EXPIRY", "form_issuing_office": "ISSUING OFFICE", "form_meningitis_expiry": "MENINGITIS EXPIRY", "form_nama_sesuai_vaksin": "NAMA SESUAI VAKSIN", "form_nomor_bpjs": "NOMOR BPJS", "form_mahram": "MAHRAM", "form_nama_ayah": "NAMA AYAH", "form_nik": "NIK", "form_alamat": "ALAMAT", "form_provinsi": "PROVINSI", "form_kabupaten": "KABUPATEN", "form_kecamatan": "KECAMATAN", "form_kelurahan": "KELURAHAN", "form_no_hp": "NO. HP", "form_status_pernikahan": "STATUS PERNIKAHAN", "form_pendidikan": "PENDIDIKAN", "form_pekerjaan": "PEKERJAAN", "form_type_of_room": "TYPE OF ROOM", "form_branch": "BRANCH"}
            for form_key, header_key in form_field_mapping_edit.items():
                form_values_for_edit[form_key] = row_data[headers_excel.index(header_key)]
            session['edit_mode_data'] = {'excel_row_num': excel_row_num, 'filename': filename, 'form_values': form_values_for_edit}
        else: flash(f"Baris {excel_row_num} tidak ditemukan di file {filename}.", "warning")
    except Exception as e: flash(f"Error memuat data edit dari {filename}: {e}", "danger")
    session['initial_tab_on_redirect'] = 'input'
    return redirect(url_for('main_page_route'))

@app.route('/update_entry', methods=['POST'])
@login_required
def update_entry_route():
    active_file = session.get('active_excel_file'); edit_data_info = session.get('edit_mode_data')
    if not active_file or not edit_data_info: flash("Sesi edit/file aktif tidak valid.", "danger"); return redirect(url_for('main_page_route'))
    row_num = edit_data_info.get('excel_row_num')
    if active_file != edit_data_info.get('filename'):
        flash("File aktif berubah saat edit. Update dibatalkan.", "danger"); session.pop('edit_mode_data', None); return redirect(url_for('main_page_route'))
    form_data_input = request.form.to_dict()
    processed_form_data, errors, age_calculation_result = process_and_validate_form(request.form)
    passport_warning = processed_form_data.pop('passport_expiry_warning', None)
    if errors:
        for error_msg in errors: flash(error_msg, 'danger')
        if passport_warning: flash(passport_warning, 'warning')
        session['edit_mode_data']['form_values'] = form_data_input 
        session['initial_tab_on_redirect'] = 'input'
        return redirect(url_for('main_page_route')) 
    if passport_warning: flash(passport_warning, 'warning')
    age_to_save = age_calculation_result
    if age_calculation_result == "N/A" and processed_form_data.get("form_date_of_birth"): age_to_save = "N/A" 
    elif not processed_form_data.get("form_date_of_birth"): age_to_save = "0"
    data_to_map = { "FULL NAME": processed_form_data.get("form_full_name"), "SEX": processed_form_data.get("form_sex"), "PLACE OF BIRTH": processed_form_data.get("form_place_of_birth"), "DATE OF BIRTH": processed_form_data.get("form_date_of_birth"), "AGE": age_to_save, "PASPORT NO": processed_form_data.get("form_pasport_no"), "DATE OF ISSUE": processed_form_data.get("form_date_of_issue"), "DATE OF EXPIRY": processed_form_data.get("form_date_of_expiry"), "ISSUING OFFICE": processed_form_data.get("form_issuing_office"), "MENINGITIS EXPIRY": processed_form_data.get("form_meningitis_expiry"), "NAMA SESUAI VAKSIN": processed_form_data.get("form_nama_sesuai_vaksin"), "NOMOR BPJS": processed_form_data.get("form_nomor_bpjs"), "MAHRAM": processed_form_data.get("form_mahram"), "NAMA AYAH": processed_form_data.get("form_nama_ayah"), "NIK": processed_form_data.get("form_nik"), "ALAMAT": processed_form_data.get("form_alamat"), "PROVINSI": processed_form_data.get("form_provinsi"), "KABUPATEN": processed_form_data.get("form_kabupaten"), "KECAMATAN": processed_form_data.get("form_kecamatan"), "KELURAHAN": processed_form_data.get("form_kelurahan"), "NO. HP": processed_form_data.get("form_no_hp"), "STATUS PERNIKAHAN": processed_form_data.get("form_status_pernikahan"), "PENDIDIKAN": processed_form_data.get("form_pendidikan"), "PEKERJAAN": processed_form_data.get("form_pekerjaan"), "TYPE OF ROOM": processed_form_data.get("form_type_of_room"), "BRANCH": processed_form_data.get("form_branch")}
    updated_row_values = [data_to_map.get(header, "") for header in headers_excel]
    try:
        filepath = get_file_path(active_file); workbook = openpyxl.load_workbook(filepath); sheet = workbook["Manifest Data"]
        if 2 <= row_num <= sheet.max_row:
            for col_idx, new_value in enumerate(updated_row_values): sheet.cell(row=row_num, column=col_idx + 1, value=new_value)
            workbook.save(filepath); flash(f'Data di file {active_file} berhasil diperbarui!', 'success')
        else: flash(f'Gagal memperbarui: Baris {row_num} tidak valid di file {active_file}.', 'danger')
    except Exception as e: flash(f'Error saat memperbarui data di file {active_file}: {e}', 'danger'); print(f"Error update Excel {active_file}: {e}")
    session.pop('edit_mode_data', None); session['initial_tab_on_redirect'] = 'input'
    per_page_val = 10
    return redirect(url_for('main_page_route', page=math.ceil(row_num / per_page_val) if per_page_val > 0 else 1, tab='preview'))

@app.route('/delete_entry/<string:filename>/<int:excel_row_num>', methods=['POST'])
@login_required
@superuser_required 
def delete_entry_route(filename, excel_row_num):
    if '/' in filename or '\\' in filename or '..' in filename: flash("Nama file tidak valid.", "danger"); return redirect(url_for('main_page_route'))
    filepath = get_file_path(filename)
    if not os.path.exists(filepath): flash(f"File '{filename}' tidak ditemukan.", 'danger'); return redirect(url_for('main_page_route'))
    try:
        workbook = openpyxl.load_workbook(filepath) 
        if "Manifest Data" not in workbook.sheetnames: flash(f'Sheet "Manifest Data" tidak ditemukan di {filename}.', 'danger'); return redirect(url_for('main_page_route'))
        sheet = workbook["Manifest Data"]; per_page_val = 10 
        page_before_delete = math.ceil((excel_row_num-1) / per_page_val) if per_page_val > 0 else 1
        if page_before_delete == 0: page_before_delete = 1
        if 2 <= excel_row_num <= sheet.max_row:
            nama_yang_dihapus = sheet.cell(row=excel_row_num, column=headers_excel.index("FULL NAME") + 1).value or "Data"
            sheet.delete_rows(excel_row_num, 1); workbook.save(filepath) 
            flash(f'{nama_yang_dihapus} dari file {filename} berhasil dihapus!', 'success')
            if sheet.max_row < 2 and filename != DEFAULT_EXCEL_FILE:
                try:
                    os.remove(filepath); flash(f"File '{filename}' dihapus karena kosong.", "info")
                    if session.get('active_excel_file') == filename: session.pop('active_excel_file', None); session['initial_tab_on_redirect'] = 'beranda'; return redirect(url_for('main_page_route')) 
                except Exception as e_remove: print(f"Gagal hapus file kosong {filename}: {e_remove}")
        else: flash(f'Gagal menghapus: Baris {excel_row_num} tidak valid di {filename}.', 'warning')
    except Exception as e: flash(f'Error saat menghapus data dari {filename}: {e}', 'danger'); print(f"Error hapus dari Excel {filename}: {e}")
    session['initial_tab_on_redirect'] = 'preview' 
    return redirect(url_for('main_page_route', page=page_before_delete, tab='preview'))

@app.route('/login', methods=['GET', 'POST'])
def login_route():
    if current_user.is_authenticated:
        return redirect(url_for('main_page_route'))
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user_obj = User.find_by_username(username)
        if user_obj:
            if user_obj.status == 'pending_approval':
                flash('Akun Anda sedang menunggu persetujuan admin.', 'warning')
            elif user_obj.status == 'active' or user_obj.status == 'superuser':
                if check_password_hash(user_obj.password, password):
                    login_user(user_obj, remember=request.form.get('remember_me') == 'y')
                    flash('Login berhasil!', 'success')
                    next_page = request.args.get('next')
                    if next_page and not (next_page.startswith('/') or next_page.startswith(request.host_url)):
                        return redirect(url_for('main_page_route'))
                    return redirect(next_page or url_for('main_page_route'))
                else:
                    flash('Username atau password salah.', 'danger')
            else: 
                flash(f'Akun Anda tidak aktif (status: {user_obj.status}). Hubungi admin.', 'danger')
        else:
            flash('Username atau password salah.', 'danger')
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register_route():
    if current_user.is_authenticated:
        return redirect(url_for('main_page_route'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        users = load_users()
        if not username or not password or not confirm_password:
            flash('Semua field registrasi wajib diisi.', 'danger')
        elif User.find_by_username(username):
            flash('Username sudah digunakan.', 'warning')
        elif password != confirm_password:
            flash('Password dan konfirmasi password tidak cocok.', 'danger')
        elif len(password) < 6: 
             flash('Password minimal harus 6 karakter.', 'danger')
        else:
            new_user_id = str(len(users) + 1) 
            while new_user_id in users: 
                new_user_id = str(int(new_user_id) + 1)
            
            user_role = 'user' 
            user_status = 'pending_approval' 
            if not users: 
                user_role = 'superuser'
                user_status = 'superuser' 

            hashed_password = generate_password_hash(password, method='pbkdf2:sha256')
            users[new_user_id] = {
                'username': username, 
                'password': hashed_password, 
                'id': new_user_id, 
                'role': user_role,
                'status': user_status 
            } 
            save_users(users)
            if user_status == 'pending_approval':
                flash('Registrasi berhasil! Akun Anda sedang menunggu persetujuan admin.', 'info')
            else: 
                flash(f'Registrasi sebagai {user_role} berhasil! Silakan login.', 'success')
            return redirect(url_for('login_route'))
    return render_template('register.html')

@app.route('/logout')
@login_required
def logout_route():
    logout_user()
    keys_to_pop = ['_user_id', '_remember', '_remember_seconds', 'active_excel_file', 'edit_mode_data', 'form_data_on_error', 'initial_tab_on_redirect','_id','csrf_token']
    for key in keys_to_pop: session.pop(key, None)
    flash('Anda telah berhasil logout.', 'info')
    return redirect(url_for('login_route'))

@app.route('/add_user_by_admin', methods=['POST'])
@login_required
@superuser_required
def add_user_by_admin_route():
    username = request.form.get('new_username', '').strip()
    password = request.form.get('new_password')
    confirm_password = request.form.get('new_confirm_password')
    role = request.form.get('new_role', 'user') 
    users = load_users()
    if not username or not password or not confirm_password:
        flash('Semua field untuk menambah pengguna wajib diisi.', 'danger')
    elif User.find_by_username(username):
        flash(f'Username "{username}" sudah digunakan.', 'warning')
    elif password != confirm_password:
        flash('Password dan konfirmasi password tidak cocok.', 'danger')
    elif len(password) < 6:
        flash('Password baru minimal harus 6 karakter.', 'danger')
    elif role not in ['user', 'superuser']:
        flash('Peran pengguna tidak valid.', 'danger')
    else:
        new_user_id = str(len(users) + 1)
        while new_user_id in users:
            new_user_id = str(int(new_user_id) + 1)
        hashed_password = generate_password_hash(password, method='pbkdf2:sha256')
        user_status_for_admin_add = 'superuser' if role == 'superuser' else 'active'
        users[new_user_id] = {'username': username, 'password': hashed_password, 'id': new_user_id, 'role': role, 'status': user_status_for_admin_add}
        save_users(users)
        flash(f'Pengguna "{username}" ({role}) berhasil ditambahkan dengan status aktif!', 'success')
    session['initial_tab_on_redirect'] = 'manage_users'
    return redirect(url_for('main_page_route'))

@app.route('/delete_user_by_admin/<string:user_id_to_delete>', methods=['POST'])
@login_required
@superuser_required
def delete_user_by_admin_route(user_id_to_delete):
    users = load_users()
    user_id_to_delete_str = str(user_id_to_delete)
    if current_user.id == user_id_to_delete_str:
        flash('Anda tidak dapat menghapus akun Anda sendiri.', 'danger')
    elif user_id_to_delete_str in users:
        superusers_count = sum(1 for u_data in users.values() if u_data.get('role') == 'superuser' and u_data.get('status') == 'superuser')
        if users[user_id_to_delete_str].get('role') == 'superuser' and users[user_id_to_delete_str].get('status') == 'superuser' and superusers_count <= 1:
            flash('Tidak dapat menghapus superuser terakhir yang aktif.', 'danger')
        else:
            deleted_username = users[user_id_to_delete_str].get('username', 'Pengguna')
            del users[user_id_to_delete_str]
            save_users(users)
            flash(f'Pengguna "{deleted_username}" berhasil dihapus.', 'success')
    else:
        flash('Pengguna tidak ditemukan atau sudah dihapus.', 'warning')
    session['initial_tab_on_redirect'] = 'manage_users'
    return redirect(url_for('main_page_route'))

@app.route('/approve_user/<string:user_id_to_approve>', methods=['POST'])
@login_required
@superuser_required
def approve_user_route(user_id_to_approve):
    users = load_users()
    user_id_str = str(user_id_to_approve)
    if user_id_str in users and users[user_id_str].get('status') == 'pending_approval':
        users[user_id_str]['status'] = 'active'
        if 'role' not in users[user_id_str] or users[user_id_str]['role'] == 'pending_approval':
            users[user_id_str]['role'] = 'user' 
        save_users(users)
        flash(f"Pengguna '{users[user_id_str]['username']}' berhasil disetujui dan diaktifkan.", "success")
    else:
        flash("Pengguna tidak ditemukan atau statusnya bukan menunggu persetujuan.", "warning")
    session['initial_tab_on_redirect'] = 'manage_users'
    return redirect(url_for('main_page_route'))

@app.route('/reject_user/<string:user_id_to_reject>', methods=['POST'])
@login_required
@superuser_required
def reject_user_route(user_id_to_reject): 
    users = load_users()
    user_id_str = str(user_id_to_reject)
    if user_id_str in users and users[user_id_str].get('status') == 'pending_approval':
        rejected_username = users[user_id_str].get('username', 'Pengguna')
        del users[user_id_str]
        save_users(users)
        flash(f"Pendaftaran pengguna '{rejected_username}' berhasil ditolak dan dihapus.", "info")
    else:
        flash("Pengguna tidak ditemukan atau statusnya bukan menunggu persetujuan.", "warning")
    session['initial_tab_on_redirect'] = 'manage_users'
    return redirect(url_for('main_page_route'))

# --- RUTE BARU UNTUK UNDUH FILE ---
@app.route('/download/<path:filename>')
@login_required # Pastikan hanya pengguna yang login bisa unduh
def download_file_route(filename):
    # Validasi nama file untuk keamanan dasar
    if ".." in filename or filename.startswith("/"):
        flash("Nama file tidak valid.", "danger")
        return redirect(url_for('main_page_route'))
    
    # Pastikan file ada di direktori penyimpanan yang diizinkan
    file_path_to_check = get_file_path(filename)
    if not os.path.exists(file_path_to_check) or not os.path.isfile(file_path_to_check):
        flash(f"File '{filename}' tidak ditemukan atau bukan file.", "danger")
        return redirect(url_for('main_page_route'))

    try:
        # send_from_directory aman digunakan untuk mengirim file dari direktori tertentu
        return send_from_directory(EXCEL_STORAGE_DIR, filename, as_attachment=True)
    except Exception as e:
        flash(f"Gagal mengunduh file '{filename}': {e}", "danger")
        print(f"Error saat download file {filename}: {e}")
        return redirect(url_for('main_page_route'))

if __name__ == '__main__':
    app.run(debug=True)

